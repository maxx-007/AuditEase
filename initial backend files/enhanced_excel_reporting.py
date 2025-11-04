#!/usr/bin/env python3
"""
Enhanced Excel Reporting Module for Compliance Audit Engine
Provides more detailed and professional Excel reports with advanced formatting
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import seaborn as sns
import io

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class EnhancedExcelReporter:
    """
    Enhanced Excel Report Generator for Compliance Audit Engine
    """
    
    def __init__(self, results, output_dir, charts_dir=None):
        """
        Initialize the Excel reporter
        
        Args:
            results: Dictionary containing compliance results
            output_dir: Directory to save the output Excel file
            charts_dir: Optional directory containing chart images
        """
        self.results = results
        self.output_dir = Path(output_dir)
        self.charts_dir = Path(charts_dir) if charts_dir else None
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create output directory if it doesn't exist
        self.output_dir.mkdir(exist_ok=True)
        
        # Define styles
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.subheader_font = Font(bold=True, color="FFFFFF")
        
        self.risk_colors = {
            'CRITICAL': 'FF0000',  # Red
            'HIGH': 'FF9900',      # Orange
            'MEDIUM': 'FFFF00',    # Yellow
            'LOW': '00FF00',       # Green
            'EXCELLENT': '00CC00'  # Dark Green
        }
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_enhanced_excel_report(self):
        """
        Generate a comprehensive Excel report with multiple detailed sheets
        
        Returns:
            Path to the generated Excel file
        """
        excel_file = self.output_dir / f"enhanced_compliance_report_{self.timestamp}.xlsx"
        
        # Create a Pandas Excel writer
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 1. Executive Summary Sheet
            self._create_executive_summary(writer)
            
            # 2. Risk Analysis Sheet
            self._create_risk_analysis(writer)
            
            # 3. Detailed Results Sheet
            self._create_detailed_results(writer)
            
            # 4. Remediation Plan Sheet
            self._create_remediation_plan(writer)
            
            # 5. Category Analysis Sheet
            self._create_category_analysis(writer)
            
            # 6. Severity Analysis Sheet
            self._create_severity_analysis(writer)
            
            # 7. Failed Rules Analysis Sheet
            self._create_failed_rules_analysis(writer)
            
            # 8. Passed Rules Analysis Sheet
            self._create_passed_rules_analysis(writer)
            
            # 9. Company Comparison Sheet
            self._create_company_comparison(writer)
            
            # 10. Gap Analysis Sheet
            self._create_gap_analysis(writer)
            
            # 11. Remediation Tracking Sheet
            self._create_remediation_tracking(writer)
        
        # Load the workbook to apply additional formatting and charts
        wb = load_workbook(excel_file)
        
        # Apply additional formatting and add charts
        self._enhance_executive_summary(wb)
        self._enhance_risk_analysis(wb)
        self._enhance_remediation_plan(wb)
        self._enhance_category_analysis(wb)
        
        # Add embedded charts if charts directory exists
        if self.charts_dir and self.charts_dir.exists():
            self._add_chart_sheet(wb)
        
        # Save the enhanced workbook
        wb.save(excel_file)
        
        logger.info(f"Enhanced Excel report saved to: {excel_file}")
        return excel_file
    
    def _create_executive_summary(self, writer):
        """Create the executive summary sheet"""
        # Prepare summary data
        summary_data = []
        
        for company, frameworks in self.results.items():
            # Overall company summary
            summary_data.append({
                'Company': company,
                'Overall_Compliance': frameworks['overall']['compliance_percentage'],
                'Overall_Risk_Level': frameworks['overall']['risk_level'],
                'Total_Critical_Issues': frameworks['overall'].get('total_critical_issues', 0),
                'Total_High_Issues': frameworks['overall'].get('total_high_issues', 0),
                'Evaluation_Date': frameworks['overall']['evaluation_date'],
                'CIS_Compliance': frameworks.get('CIS', {}).get('compliance_percentage', 0),
                'CIS_Risk': frameworks.get('CIS', {}).get('risk_level', 'N/A'),
                'ISO27001_Compliance': frameworks.get('ISO27001', {}).get('compliance_percentage', 0),
                'ISO27001_Risk': frameworks.get('ISO27001', {}).get('risk_level', 'N/A'),
                'RBI_Compliance': frameworks.get('RBI', {}).get('compliance_percentage', 0),
                'RBI_Risk': frameworks.get('RBI', {}).get('risk_level', 'N/A')
            })
        
        # Convert to DataFrame and write to Excel
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Executive_Summary', index=False)
    
    def _create_risk_analysis(self, writer):
        """Create the risk analysis sheet"""
        risk_data = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    risk_data.append({
                        'Company': company,
                        'Framework': framework,
                        'Compliance_Percentage': metrics['compliance_percentage'],
                        'Risk_Level': metrics['risk_level'],
                        'Total_Rules': metrics['total_rules'],
                        'Passed_Rules': metrics['passed_rules'],
                        'Failed_Rules': metrics['failed_rules'],
                        'Missing_Data_Rules': metrics['missing_data_rules'],
                        'Error_Rules': metrics['error_rules'],
                        'Critical_Severity_Count': metrics['severity_breakdown'].get('CRITICAL', 0),
                        'High_Severity_Count': metrics['severity_breakdown'].get('HIGH', 0),
                        'Medium_Severity_Count': metrics['severity_breakdown'].get('MEDIUM', 0),
                        'Low_Severity_Count': metrics['severity_breakdown'].get('LOW', 0),
                        'Pass_Rate': metrics['passed_rules'] / metrics['total_rules'] * 100 if metrics['total_rules'] > 0 else 0
                    })
        
        df_risk = pd.DataFrame(risk_data)
        df_risk.to_excel(writer, sheet_name='Risk_Analysis', index=False)
    
    def _create_detailed_results(self, writer):
        """Create the detailed results sheet"""
        detailed_data = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    for rule in metrics['rule_details']:
                        # Basic rule data
                        rule_data = {
                            'Company': company,
                            'Framework': framework,
                            'Rule_ID': rule.get('rule_id', 'N/A'),
                            'Category': rule.get('category', 'General'),
                            'Description': rule.get('description', 'N/A'),
                            'Status': rule.get('status', 'N/A'),
                            'Severity': rule.get('severity', 'MEDIUM'),
                            'Weight': rule.get('weight', 1),
                            'Score': rule.get('score', 0),
                            'Field_Checked': rule.get('field', 'N/A'),
                            'Expected_Value': rule.get('expected_value', 'N/A'),
                            'Actual_Value': rule.get('actual_value', 'N/A'),
                            'Status_Message': rule.get('message', 'N/A')
                        }
                        
                        # Enhanced remediation data
                        if 'remediation' in rule and isinstance(rule['remediation'], dict):
                            rem = rule['remediation']
                            rule_data.update({
                                'Remediation_Description': rem.get('description', ''),
                                'Impact_Analysis': rem.get('impact', ''),
                                'Priority_Level': rem.get('priority', ''),
                                'Effort_Estimate': rem.get('effort', ''),
                                'Business_Justification': rem.get('business_justification', ''),
                                'Has_Automation_Scripts': 'Yes' if rem.get('scripts') else 'No',
                                'Script_Count': len(rem.get('scripts', [])),
                                'References': '; '.join(rem.get('references', []))
                            })
                        else:
                            rule_data.update({
                                'Remediation_Description': rule.get('remediation', ''),
                                'Impact_Analysis': '',
                                'Priority_Level': '',
                                'Effort_Estimate': '',
                                'Business_Justification': '',
                                'Has_Automation_Scripts': 'No',
                                'Script_Count': 0,
                                'References': ''
                            })
                        
                        detailed_data.append(rule_data)
        
        df_detailed = pd.DataFrame(detailed_data)
        df_detailed.to_excel(writer, sheet_name='Detailed_Results', index=False)
    
    def _create_remediation_plan(self, writer):
        """Create the remediation plan sheet"""
        remediation_data = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    for rule in metrics['rule_details']:
                        if rule['status'] in ['FAIL', 'MISSING_DATA', 'ERROR']:
                            # Calculate priority score
                            severity_scores = {'CRITICAL': 100, 'HIGH': 75, 'MEDIUM': 50, 'LOW': 25}
                            weight_multiplier = rule.get('weight', 1)
                            severity = rule.get('severity', 'MEDIUM')
                            base_score = severity_scores.get(severity, 50)
                            priority_score = base_score + (weight_multiplier * 5)
                            
                            # Get remediation timeline
                            timeline_map = {
                                'CRITICAL': 'Immediate (24 hours)',
                                'HIGH': 'Urgent (1 week)',
                                'MEDIUM': 'Standard (1 month)',
                                'LOW': 'Planned (3 months)'
                            }
                            timeline = timeline_map.get(severity, 'Standard (1 month)')
                            
                            # Get recommended team
                            category = rule.get('category', 'General')
                            team_map = {
                                'Access Control': 'Security Team',
                                'Asset Management': 'IT Operations',
                                'Configuration Management': 'System Administration',
                                'Network Security': 'Network Team',
                                'Logging and Monitoring': 'SOC Team',
                                'Vulnerability Management': 'Security Team',
                                'Incident Response': 'Security Team',
                                'Business Continuity': 'Risk Management'
                            }
                            team = team_map.get(category, 'IT Team')
                            
                            # Get cost category
                            if severity in ['CRITICAL', 'HIGH']:
                                cost = 'High Priority ($$)'
                            elif category in ['Asset Management', 'Logging and Monitoring', 'Business Continuity']:
                                cost = 'Medium-High ($-$$)'
                            elif category in ['Configuration Management']:
                                cost = 'Low ($)'
                            else:
                                cost = 'Medium ($)'
                            
                            # Get remediation info
                            if 'remediation' in rule and isinstance(rule['remediation'], dict):
                                rem = rule['remediation']
                                remediation_desc = rem.get('description', '')
                                impact = rem.get('impact', '')
                                effort = rem.get('effort', '')
                                business_justification = rem.get('business_justification', '')
                                has_scripts = 'Yes' if rem.get('scripts') else 'No'
                            else:
                                remediation_desc = rule.get('remediation', '')
                                impact = ''
                                effort = ''
                                business_justification = ''
                                has_scripts = 'No'
                            
                            remediation_entry = {
                                'Company': company,
                                'Framework': framework,
                                'Rule_ID': rule.get('rule_id', 'N/A'),
                                'Category': category,
                                'Description': rule.get('description', 'N/A'),
                                'Current_Status': rule['status'],
                                'Severity': severity,
                                'Priority_Score': priority_score,
                                'Remediation_Description': remediation_desc,
                                'Impact_Analysis': impact,
                                'Effort_Estimate': effort,
                                'Business_Justification': business_justification,
                                'Recommended_Timeline': timeline,
                                'Assigned_Team': team,
                                'Cost_Category': cost,
                                'Compliance_Frameworks_Affected': framework,
                                'Has_Automation_Scripts': has_scripts
                            }
                            
                            remediation_data.append(remediation_entry)
        
        # Sort by priority score (descending)
        df_remediation = pd.DataFrame(remediation_data)
        if not df_remediation.empty:
            df_remediation = df_remediation.sort_values(['Priority_Score', 'Company'], ascending=[False, True])
        
        df_remediation.to_excel(writer, sheet_name='Remediation_Plan', index=False)
    
    def _create_category_analysis(self, writer):
        """Create the category analysis sheet"""
        category_data = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    for category, stats in metrics.get('category_breakdown', {}).items():
                        category_data.append({
                            'Company': company,
                            'Framework': framework,
                            'Category': category,
                            'Total_Rules': stats['total'],
                            'Passed_Rules': stats['passed'],
                            'Failed_Rules': stats['failed'],
                            'Compliance_Percentage': stats['compliance_pct'],
                            'Risk_Assessment': self._get_risk_level(stats['compliance_pct'])
                        })
        
        df_category = pd.DataFrame(category_data)
        df_category.to_excel(writer, sheet_name='Category_Analysis', index=False)
    
    def _create_severity_analysis(self, writer):
        """Create the severity analysis sheet"""
        severity_data = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    for severity, count in metrics['severity_breakdown'].items():
                        if count > 0:
                            severity_data.append({
                                'Company': company,
                                'Framework': framework,
                                'Severity': severity,
                                'Count': count,
                                'Percentage_of_Total': round((count / metrics['total_rules']) * 100, 2)
                            })
        
        df_severity = pd.DataFrame(severity_data)
        df_severity.to_excel(writer, sheet_name='Severity_Analysis', index=False)
    
    def _create_failed_rules_analysis(self, writer):
        """Create the failed rules analysis sheet"""
        failed_rules = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    for rule in metrics['rule_details']:
                        if rule['status'] in ['FAIL', 'MISSING_DATA', 'ERROR']:
                            failed_rule = {
                                'Company': company,
                                'Framework': framework,
                                'Rule_ID': rule.get('rule_id', 'N/A'),
                                'Category': rule.get('category', 'General'),
                                'Description': rule.get('description', 'N/A'),
                                'Status': rule['status'],
                                'Severity': rule.get('severity', 'MEDIUM'),
                                'Field_Checked': rule.get('field', 'N/A'),
                                'Expected_Value': rule.get('expected_value', 'N/A'),
                                'Actual_Value': rule.get('actual_value', 'N/A'),
                                'Status_Message': rule.get('message', 'N/A'),
                                'Remediation': rule.get('remediation', {}).get('description', '') if isinstance(rule.get('remediation'), dict) else rule.get('remediation', '')
                            }
                            
                            failed_rules.append(failed_rule)
        
        df_failed = pd.DataFrame(failed_rules)
        df_failed.to_excel(writer, sheet_name='Failed_Rules_Analysis', index=False)
    
    def _create_passed_rules_analysis(self, writer):
        """Create the passed rules analysis sheet"""
        passed_rules = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    for rule in metrics['rule_details']:
                        if rule['status'] == 'PASS':
                            passed_rule = {
                                'Company': company,
                                'Framework': framework,
                                'Rule_ID': rule.get('rule_id', 'N/A'),
                                'Category': rule.get('category', 'General'),
                                'Description': rule.get('description', 'N/A'),
                                'Severity': rule.get('severity', 'MEDIUM'),
                                'Field_Checked': rule.get('field', 'N/A'),
                                'Expected_Value': rule.get('expected_value', 'N/A'),
                                'Actual_Value': rule.get('actual_value', 'N/A')
                            }
                            
                            passed_rules.append(passed_rule)
        
        df_passed = pd.DataFrame(passed_rules)
        df_passed.to_excel(writer, sheet_name='Passed_Rules_Analysis', index=False)
    
    def _create_company_comparison(self, writer):
        """Create the company comparison sheet"""
        comparison_data = []
        
        for framework in set(fw for comp in self.results.values() for fw in comp.keys() if fw != 'overall'):
            for company in self.results.keys():
                if framework in self.results[company]:
                    metrics = self.results[company][framework]
                    comparison_data.append({
                        'Framework': framework,
                        'Company': company,
                        'Compliance_Pct': metrics['compliance_percentage'],
                        'Risk_Level': metrics['risk_level'],
                        'Passed_Rules': metrics['passed_rules'],
                        'Total_Rules': metrics['total_rules'],
                        'Pass_Rate': round((metrics['passed_rules'] / metrics['total_rules']) * 100, 2) if metrics['total_rules'] > 0 else 0
                    })
        
        df_comparison = pd.DataFrame(comparison_data)
        df_comparison.to_excel(writer, sheet_name='Company_Comparison', index=False)
    
    def _create_gap_analysis(self, writer):
        """Create the gap analysis sheet"""
        gap_data = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    for category, stats in metrics.get('category_breakdown', {}).items():
                        current = stats.get('compliance_pct', 0)
                        target = 85  # Default target (LOW risk threshold)
                        gap = target - current
                        
                        if gap > 0:  # Only include categories with gaps
                            gap_data.append({
                                'Company': company,
                                'Framework': framework,
                                'Category': category,
                                'Current_Compliance': current,
                                'Target_Compliance': target,
                                'Gap': gap,
                                'Gap_Percentage': round((gap / target) * 100, 2),
                                'Priority': 'High' if gap > 50 else 'Medium' if gap > 25 else 'Low',
                                'Estimated_Effort': 'High (1-2 months)' if gap > 50 else 'Medium (2-4 weeks)' if gap > 25 else 'Low (1-2 weeks)'
                            })
        
        df_gap = pd.DataFrame(gap_data)
        df_gap = df_gap.sort_values(['Gap', 'Company'], ascending=[False, True])
        df_gap.to_excel(writer, sheet_name='Gap_Analysis', index=False)
    
    def _create_remediation_tracking(self, writer):
        """Create the remediation tracking sheet"""
        tracking_data = []
        
        for company, frameworks in self.results.items():
            for framework, metrics in frameworks.items():
                if framework != 'overall':
                    for rule in metrics['rule_details']:
                        if rule['status'] in ['FAIL', 'MISSING_DATA', 'ERROR']:
                            # Calculate due date based on severity
                            severity = rule.get('severity', 'MEDIUM')
                            if severity == 'CRITICAL':
                                due_days = 7
                            elif severity == 'HIGH':
                                due_days = 30
                            elif severity == 'MEDIUM':
                                due_days = 60
                            else:
                                due_days = 90
                            
                            due_date = (datetime.now() + timedelta(days=due_days)).strftime("%Y-%m-%d")
                            
                            # Add verification method based on rule type
                            category = rule.get('category', 'General')
                            if 'Configuration' in category:
                                verify = "Configuration Review"
                            elif 'Network' in category:
                                verify = "Network Scan"
                            elif 'Access' in category:
                                verify = "Access Control Audit"
                            else:
                                verify = "Manual Verification"
                            
                            tracking_entry = {
                                'Company': company,
                                'Framework': framework,
                                'Rule_ID': rule.get('rule_id', 'N/A'),
                                'Description': rule.get('description', 'N/A'),
                                'Status': 'Open',
                                'Assigned_To': '',
                                'Due_Date': due_date,
                                'Progress': '0%',
                                'Notes': '',
                                'Verification_Method': verify
                            }
                            
                            tracking_data.append(tracking_entry)
        
        df_tracking = pd.DataFrame(tracking_data)
        df_tracking.to_excel(writer, sheet_name='Remediation_Tracking', index=False)
    
    def _enhance_executive_summary(self, wb):
        """Enhance the executive summary sheet with formatting and charts"""
        ws = wb['Executive_Summary']
        
        # Apply header formatting
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.border
            
            # Adjust column width
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Add title
        ws.insert_rows(1, 3)
        ws.merge_cells('A1:L1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "COMPLIANCE AUDIT EXECUTIVE SUMMARY"
        title_cell.font = Font(size=16, bold=True, color="1F4E79")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add subtitle with date
        ws.merge_cells('A2:L2')
        subtitle_cell = ws.cell(row=2, column=2)
        subtitle_cell.value = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        subtitle_cell.font = Font(size=12, italic=True)
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add spacer row
        ws.row_dimensions[3].height = 10
        
        # Apply conditional formatting to risk levels
        for row in range(5, ws.max_row + 1):  # Start from row 5 (after headers)
            risk_cell = ws.cell(row=row, column=3)  # Overall_Risk_Level column
            if risk_cell.value in self.risk_colors:
                risk_cell.fill = PatternFill(start_color=self.risk_colors[risk_cell.value], 
                                            end_color=self.risk_colors[risk_cell.value], 
                                            fill_type="solid")
                risk_cell.font = Font(bold=True, color="000000" if risk_cell.value in ['LOW', 'MEDIUM', 'EXCELLENT'] else "FFFFFF")
            
            # Format compliance percentages
            for col in [2, 7, 9, 11]:  # Compliance percentage columns
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.value = f"{cell.value:.1f}%"
                    cell.alignment = Alignment(horizontal='center')
    
    def _enhance_risk_analysis(self, wb):
        """Enhance the risk analysis sheet with formatting and charts"""
        ws = wb['Risk_Analysis']
        
        # Apply header formatting
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.border
            
            # Adjust column width
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Apply conditional formatting to risk levels
        for row in range(2, ws.max_row + 1):
            risk_cell = ws.cell(row=row, column=4)  # Risk_Level column
            if risk_cell.value in self.risk_colors:
                risk_cell.fill = PatternFill(start_color=self.risk_colors[risk_cell.value], 
                                            end_color=self.risk_colors[risk_cell.value], 
                                            fill_type="solid")
                risk_cell.font = Font(bold=True, color="000000" if risk_cell.value in ['LOW', 'MEDIUM', 'EXCELLENT'] else "FFFFFF")
    
    def _enhance_remediation_plan(self, wb):
        """Enhance the remediation plan sheet with formatting and prioritization"""
        ws = wb['Remediation_Plan']
        
        # Apply header formatting
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.border
            
            # Adjust column width
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _enhance_category_analysis(self, wb):
        """Enhance the category analysis sheet with formatting and charts"""
        ws = wb['Category_Analysis']
        
        # Apply header formatting
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.border
            
            # Adjust column width
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _add_chart_sheet(self, wb):
        """Add a sheet with embedded charts"""
        # Create a new sheet for charts
        if 'Charts' in wb.sheetnames:
            ws = wb['Charts']
        else:
            ws = wb.create_sheet('Charts')
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "COMPLIANCE VISUALIZATION CHARTS"
        title_cell.font = Font(size=16, bold=True, color="1F4E79")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add charts from the charts directory
        chart_files = list(self.charts_dir.glob('*.png'))
        
        row = 3
        for i, chart_file in enumerate(chart_files):
            try:
                img = Image(chart_file)
                # Scale image to fit in Excel
                img.width = 600
                img.height = 400
                
                # Add chart title
                ws.cell(row=row, column=1).value = chart_file.stem.replace('_', ' ').title()
                ws.cell(row=row, column=1).font = Font(size=12, bold=True)
                row += 1
                
                # Add image
                ws.add_image(img, f'B{row}')
                row += 25  # Space for the image
                
                # Add spacer
                row += 2
            except Exception as e:
                logger.error(f"Error adding chart {chart_file}: {e}")
    
    def _get_risk_level(self, compliance_pct):
        """Get risk level based on compliance percentage"""
        if compliance_pct >= 90:
            return 'LOW'
        elif compliance_pct >= 75:
            return 'MEDIUM'
        elif compliance_pct >= 50:
            return 'HIGH'
        else:
            return 'CRITICAL'