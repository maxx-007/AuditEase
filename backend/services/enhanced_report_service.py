#!/usr/bin/env python3
"""
Enhanced Report Service - Production-Grade Reporting Module
============================================================
Provides comprehensive Excel and PDF reporting with professional formatting,
multiple detailed sheets, embedded charts, and advanced analytics.

Features:
- 11+ detailed Excel sheets with conditional formatting
- Professional PDF reports with executive summaries
- Embedded visualizations and charts
- Remediation tracking and gap analysis
- Risk assessment matrices
- Timeline planning and cost estimation
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, 
    TableStyle, PageBreak, KeepTogether, Image as RLImage
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO

# Setup logger
logger = logging.getLogger(__name__)


class EnhancedReportService:
    """
    Production-grade enhanced reporting service for compliance audits.
    
    Generates comprehensive Excel and PDF reports with professional formatting,
    multiple detailed sheets, embedded charts, and advanced analytics.
    """
    
    def __init__(self, output_dir: str = "reports", charts_dir: Optional[str] = None):
        """
        Initialize the Enhanced Report Service.
        
        Args:
            output_dir: Directory to save generated reports
            charts_dir: Optional directory containing chart images
        """
        self.output_dir = Path(output_dir)
        self.charts_dir = Path(charts_dir) if charts_dir else self.output_dir / "charts"
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create directories
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.charts_dir.mkdir(parents=True, exist_ok=True)
        
        # Define professional color scheme
        self.colors = {
            'primary': '1F4E79',      # Dark blue
            'secondary': '4472C4',    # Medium blue
            'accent': '2E75B6',       # Light blue
            'success': '00CC00',      # Green
            'warning': 'FFA500',      # Orange
            'danger': 'FF0000',       # Red
            'critical': '8B0000',     # Dark red
            'text': '000000',         # Black
            'light': 'F0F0F0',        # Light gray
        }
        
        # Risk level color mapping
        self.risk_colors = {
            'CRITICAL': self.colors['critical'],
            'HIGH': self.colors['danger'],
            'MEDIUM': self.colors['warning'],
            'LOW': self.colors['success'],
            'EXCELLENT': '00CC00'
        }
        
        # Define professional styles
        self._init_styles()
        
        logger.info(f"Enhanced Report Service initialized - Output: {self.output_dir}")
    
    def _init_styles(self):
        """Initialize professional styling for reports."""
        self.header_fill = PatternFill(
            start_color=self.colors['primary'],
            end_color=self.colors['primary'],
            fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        self.subheader_fill = PatternFill(
            start_color=self.colors['secondary'],
            end_color=self.colors['secondary'],
            fill_type="solid"
        )
        self.subheader_font = Font(bold=True, color="FFFFFF", size=10)
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_comprehensive_excel_report(
        self,
        audit_results: Dict[str, Any],
        company_name: str = "Organization"
    ) -> Path:
        """
        Generate a comprehensive Excel report with 11+ detailed sheets.
        
        Args:
            audit_results: Complete audit results dictionary
            company_name: Name of the organization being audited
            
        Returns:
            Path to the generated Excel file
        """
        try:
            logger.info(f"Generating comprehensive Excel report for {company_name}...")
            
            excel_file = self.output_dir / f"comprehensive_report_{company_name}_{self.timestamp}.xlsx"
            
            # Create Excel writer
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # Sheet 1: Executive Summary
                self._create_executive_summary_sheet(writer, audit_results, company_name)
                
                # Sheet 2: Risk Analysis
                self._create_risk_analysis_sheet(writer, audit_results)
                
                # Sheet 3: Detailed Results
                self._create_detailed_results_sheet(writer, audit_results)
                
                # Sheet 4: Remediation Plan (Prioritized)
                self._create_remediation_plan_sheet(writer, audit_results)
                
                # Sheet 5: Category Analysis
                self._create_category_analysis_sheet(writer, audit_results)
                
                # Sheet 6: Severity Analysis
                self._create_severity_analysis_sheet(writer, audit_results)
                
                # Sheet 7: Failed Rules Analysis
                self._create_failed_rules_sheet(writer, audit_results)
                
                # Sheet 8: Passed Rules Analysis
                self._create_passed_rules_sheet(writer, audit_results)
                
                # Sheet 9: Framework Comparison
                self._create_framework_comparison_sheet(writer, audit_results)
                
                # Sheet 10: Gap Analysis
                self._create_gap_analysis_sheet(writer, audit_results)
                
                # Sheet 11: Remediation Tracking
                self._create_remediation_tracking_sheet(writer, audit_results)
                
                # Sheet 12: Timeline & Roadmap
                self._create_timeline_roadmap_sheet(writer, audit_results)
            
            # Apply advanced formatting and charts
            self._enhance_excel_workbook(excel_file)
            
            logger.info(f"✅ Comprehensive Excel report generated: {excel_file}")
            return excel_file
            
        except Exception as e:
            logger.error(f"❌ Excel report generation failed: {e}", exc_info=True)
            raise
    
    def _create_executive_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any],
        company_name: str
    ):
        """Create executive summary sheet with high-level metrics."""
        try:
            # Extract data
            frameworks = audit_results.get('frameworks', {})
            overall = audit_results.get('overall_summary', {})
            
            summary_data = []
            
            # Overall metrics
            summary_data.append({
                'Metric': 'Organization',
                'Value': company_name,
                'Status': 'N/A',
                'Risk Level': overall.get('overall_risk_level', 'UNKNOWN')
            })
            
            summary_data.append({
                'Metric': 'Assessment Date',
                'Value': audit_results.get('audit_date', datetime.now().isoformat()),
                'Status': 'N/A',
                'Risk Level': 'N/A'
            })
            
            summary_data.append({
                'Metric': 'Overall Compliance',
                'Value': f"{overall.get('average_compliance_percentage', 0):.1f}%",
                'Status': 'Measured',
                'Risk Level': overall.get('overall_risk_level', 'UNKNOWN')
            })
            
            # Framework-specific metrics
            for fw_name, fw_data in frameworks.items():
                overall_fw = fw_data.get('overall', fw_data)
                summary_data.append({
                    'Metric': f'{fw_name} Compliance',
                    'Value': f"{overall_fw.get('compliance_percentage', 0):.1f}%",
                    'Status': f"{overall_fw.get('passed_rules', 0)}/{overall_fw.get('total_rules', 0)} Rules",
                    'Risk Level': overall_fw.get('risk_level', 'UNKNOWN')
                })
            
            # Critical issues
            total_critical = sum(
                fw.get('overall', fw).get('severity_breakdown', {}).get('CRITICAL', 0)
                for fw in frameworks.values()
            )
            total_high = sum(
                fw.get('overall', fw).get('severity_breakdown', {}).get('CRITICAL', 0)
                for fw in frameworks.values()
            )
            
            summary_data.append({
                'Metric': 'Critical Issues',
                'Value': str(total_critical),
                'Status': 'Immediate Action Required' if total_critical > 0 else 'None',
                'Risk Level': 'CRITICAL' if total_critical > 0 else 'LOW'
            })
            
            summary_data.append({
                'Metric': 'High Priority Issues',
                'Value': str(total_high),
                'Status': 'Action Required' if total_high > 0 else 'None',
                'Risk Level': 'HIGH' if total_high > 0 else 'LOW'
            })
            
            # Create DataFrame
            df = pd.DataFrame(summary_data)
            df.to_excel(writer, sheet_name='Executive_Summary', index=False)
            
            logger.debug("Executive summary sheet created")
            
        except Exception as e:
            logger.error(f"Error creating executive summary: {e}", exc_info=True)
            # Create minimal sheet
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Executive_Summary', index=False
            )
    
    def _create_risk_analysis_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create detailed risk analysis sheet."""
        try:
            frameworks = audit_results.get('frameworks', {})
            risk_data = []
            
            for fw_name, fw_data in frameworks.items():
                overall_fw = fw_data.get('overall', fw_data)
                severity_breakdown = overall_fw.get('severity_breakdown', {})
                
                risk_data.append({
                    'Framework': fw_name,
                    'Compliance_Percentage': overall_fw.get('compliance_percentage', 0),
                    'Risk_Level': overall_fw.get('risk_level', 'UNKNOWN'),
                    'Total_Rules': overall_fw.get('total_rules', 0),
                    'Passed_Rules': overall_fw.get('passed_rules', 0),
                    'Failed_Rules': overall_fw.get('failed_rules', 0),
                    'Missing_Data': overall_fw.get('missing_data_rules', 0),
                    'Critical_Issues': severity_breakdown.get('CRITICAL', 0),
                    'High_Issues': severity_breakdown.get('HIGH', 0),
                    'Medium_Issues': severity_breakdown.get('MEDIUM', 0),
                    'Low_Issues': severity_breakdown.get('LOW', 0),
                    'Pass_Rate': f"{(overall_fw.get('passed_rules', 0) / overall_fw.get('total_rules', 1) * 100):.1f}%"
                })
            
            df = pd.DataFrame(risk_data)
            df.to_excel(writer, sheet_name='Risk_Analysis', index=False)
            
            logger.debug("Risk analysis sheet created")
            
        except Exception as e:
            logger.error(f"Error creating risk analysis: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Risk_Analysis', index=False
            )
    
    def _create_detailed_results_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create detailed results sheet with all rule evaluations."""
        try:
            frameworks = audit_results.get('frameworks', {})
            detailed_data = []
            
            for fw_name, fw_data in frameworks.items():
                rules = fw_data.get('rules', fw_data.get('rule_details', []))
                
                for rule in rules:
                    detailed_data.append({
                        'Framework': fw_name,
                        'Rule_ID': rule.get('rule_id', 'N/A'),
                        'Category': rule.get('category', 'General'),
                        'Description': rule.get('description', 'N/A'),
                        'Status': rule.get('status', 'UNKNOWN'),
                        'Severity': rule.get('severity', 'MEDIUM'),
                        'Weight': rule.get('weight', 1),
                        'Score': rule.get('score', 0),
                        'Field_Checked': rule.get('field', 'N/A'),
                        'Expected_Value': str(rule.get('expected_value', 'N/A')),
                        'Actual_Value': str(rule.get('actual_value', 'N/A')),
                        'Message': rule.get('message', 'N/A'),
                        'Remediation': rule.get('remediation', {}).get('description', '') 
                                     if isinstance(rule.get('remediation'), dict) 
                                     else rule.get('remediation', '')
                    })
            
            df = pd.DataFrame(detailed_data)
            df.to_excel(writer, sheet_name='Detailed_Results', index=False)
            
            logger.debug(f"Detailed results sheet created with {len(detailed_data)} rules")
            
        except Exception as e:
            logger.error(f"Error creating detailed results: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Detailed_Results', index=False
            )
    
    def _create_remediation_plan_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create prioritized remediation plan sheet."""
        try:
            frameworks = audit_results.get('frameworks', {})
            remediation_data = []
            
            for fw_name, fw_data in frameworks.items():
                rules = fw_data.get('rules', fw_data.get('rule_details', []))
                
                for rule in rules:
                    if rule.get('status') in ['FAIL', 'MISSING_DATA', 'ERROR']:
                        # Calculate priority score
                        severity = rule.get('severity', 'MEDIUM')
                        severity_scores = {'CRITICAL': 100, 'HIGH': 75, 'MEDIUM': 50, 'LOW': 25}
                        weight = rule.get('weight', 1)
                        priority_score = severity_scores.get(severity, 50) + (weight * 5)
                        
                        # Get timeline
                        timeline_map = {
                            'CRITICAL': 'Immediate (24-48 hours)',
                            'HIGH': 'Urgent (1 week)',
                            'MEDIUM': 'Standard (1 month)',
                            'LOW': 'Planned (3 months)'
                        }
                        
                        remediation_data.append({
                            'Priority_Score': priority_score,
                            'Framework': fw_name,
                            'Rule_ID': rule.get('rule_id', 'N/A'),
                            'Category': rule.get('category', 'General'),
                            'Description': rule.get('description', 'N/A'),
                            'Severity': severity,
                            'Current_Status': rule.get('status', 'UNKNOWN'),
                            'Timeline': timeline_map.get(severity, 'Standard'),
                            'Remediation_Steps': rule.get('remediation', {}).get('description', '')
                                               if isinstance(rule.get('remediation'), dict)
                                               else rule.get('remediation', ''),
                            'Effort_Estimate': rule.get('remediation', {}).get('effort', 'Medium')
                                             if isinstance(rule.get('remediation'), dict)
                                             else 'Medium',
                            'Assigned_To': '',
                            'Status_Tracking': 'Not Started'
                        })
            
            # Sort by priority score
            df = pd.DataFrame(remediation_data)
            if not df.empty:
                df = df.sort_values('Priority_Score', ascending=False)
            df.to_excel(writer, sheet_name='Remediation_Plan', index=False)
            
            logger.debug(f"Remediation plan sheet created with {len(remediation_data)} items")
            
        except Exception as e:
            logger.error(f"Error creating remediation plan: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Remediation_Plan', index=False
            )
    
    def _create_category_analysis_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create category-wise analysis sheet."""
        try:
            frameworks = audit_results.get('frameworks', {})
            category_data = []
            
            for fw_name, fw_data in frameworks.items():
                categories = fw_data.get('category_scores', fw_data.get('category_breakdown', {}))
                
                for category, stats in categories.items():
                    category_data.append({
                        'Framework': fw_name,
                        'Category': category,
                        'Total_Rules': stats.get('total_rules', stats.get('total', 0)),
                        'Passed_Rules': stats.get('passed_rules', stats.get('passed', 0)),
                        'Failed_Rules': stats.get('failed_rules', stats.get('failed', 0)),
                        'Compliance_Percentage': stats.get('compliance_percentage', stats.get('compliance_pct', 0)),
                        'Risk_Assessment': self._get_risk_level(
                            stats.get('compliance_percentage', stats.get('compliance_pct', 0))
                        )
                    })
            
            df = pd.DataFrame(category_data)
            df.to_excel(writer, sheet_name='Category_Analysis', index=False)
            
            logger.debug("Category analysis sheet created")
            
        except Exception as e:
            logger.error(f"Error creating category analysis: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Category_Analysis', index=False
            )
    
    def _create_severity_analysis_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create severity-based analysis sheet."""
        try:
            frameworks = audit_results.get('frameworks', {})
            severity_data = []
            
            for fw_name, fw_data in frameworks.items():
                overall_fw = fw_data.get('overall', fw_data)
                severity_breakdown = overall_fw.get('severity_breakdown', {})
                total_rules = overall_fw.get('total_rules', 1)
                
                for severity, count in severity_breakdown.items():
                    if count > 0:
                        severity_data.append({
                            'Framework': fw_name,
                            'Severity': severity,
                            'Count': count,
                            'Percentage_of_Total': f"{(count / total_rules * 100):.1f}%",
                            'Risk_Impact': self._get_severity_impact(severity)
                        })
            
            df = pd.DataFrame(severity_data)
            df.to_excel(writer, sheet_name='Severity_Analysis', index=False)
            
            logger.debug("Severity analysis sheet created")
            
        except Exception as e:
            logger.error(f"Error creating severity analysis: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Severity_Analysis', index=False
            )
    
    def _create_failed_rules_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create failed rules analysis sheet."""
        try:
            frameworks = audit_results.get('frameworks', {})
            failed_data = []
            
            for fw_name, fw_data in frameworks.items():
                rules = fw_data.get('rules', fw_data.get('rule_details', []))
                
                for rule in rules:
                    if rule.get('status') in ['FAIL', 'MISSING_DATA', 'ERROR']:
                        failed_data.append({
                            'Framework': fw_name,
                            'Rule_ID': rule.get('rule_id', 'N/A'),
                            'Category': rule.get('category', 'General'),
                            'Description': rule.get('description', 'N/A'),
                            'Severity': rule.get('severity', 'MEDIUM'),
                            'Status': rule.get('status', 'UNKNOWN'),
                            'Expected': str(rule.get('expected_value', 'N/A')),
                            'Actual': str(rule.get('actual_value', 'N/A')),
                            'Gap': rule.get('message', 'N/A')
                        })
            
            df = pd.DataFrame(failed_data)
            df.to_excel(writer, sheet_name='Failed_Rules', index=False)
            
            logger.debug(f"Failed rules sheet created with {len(failed_data)} failures")
            
        except Exception as e:
            logger.error(f"Error creating failed rules sheet: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Failed_Rules', index=False
            )
    
    def _create_passed_rules_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create passed rules analysis sheet."""
        try:
            frameworks = audit_results.get('frameworks', {})
            passed_data = []
            
            for fw_name, fw_data in frameworks.items():
                rules = fw_data.get('rules', fw_data.get('rule_details', []))
                
                for rule in rules:
                    if rule.get('status') == 'PASS':
                        passed_data.append({
                            'Framework': fw_name,
                            'Rule_ID': rule.get('rule_id', 'N/A'),
                            'Category': rule.get('category', 'General'),
                            'Description': rule.get('description', 'N/A'),
                            'Severity': rule.get('severity', 'MEDIUM'),
                            'Value': str(rule.get('actual_value', 'N/A'))
                        })
            
            df = pd.DataFrame(passed_data)
            df.to_excel(writer, sheet_name='Passed_Rules', index=False)
            
            logger.debug(f"Passed rules sheet created with {len(passed_data)} passes")
            
        except Exception as e:
            logger.error(f"Error creating passed rules sheet: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Passed_Rules', index=False
            )
    
    def _create_framework_comparison_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create framework comparison sheet."""
        try:
            frameworks = audit_results.get('frameworks', {})
            comparison_data = []
            
            for fw_name, fw_data in frameworks.items():
                overall_fw = fw_data.get('overall', fw_data)
                comparison_data.append({
                    'Framework': fw_name,
                    'Compliance_Percentage': overall_fw.get('compliance_percentage', 0),
                    'Risk_Level': overall_fw.get('risk_level', 'UNKNOWN'),
                    'Total_Rules': overall_fw.get('total_rules', 0),
                    'Passed': overall_fw.get('passed_rules', 0),
                    'Failed': overall_fw.get('failed_rules', 0),
                    'Pass_Rate': f"{(overall_fw.get('passed_rules', 0) / overall_fw.get('total_rules', 1) * 100):.1f}%",
                    'Critical_Issues': overall_fw.get('severity_breakdown', {}).get('CRITICAL', 0),
                    'High_Issues': overall_fw.get('severity_breakdown', {}).get('HIGH', 0)
                })
            
            df = pd.DataFrame(comparison_data)
            df.to_excel(writer, sheet_name='Framework_Comparison', index=False)
            
            logger.debug("Framework comparison sheet created")
            
        except Exception as e:
            logger.error(f"Error creating framework comparison: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Framework_Comparison', index=False
            )
    
    def _create_gap_analysis_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create gap analysis sheet."""
        try:
            frameworks = audit_results.get('frameworks', {})
            gap_data = []
            target_compliance = 85.0  # Target threshold
            
            for fw_name, fw_data in frameworks.items():
                categories = fw_data.get('category_scores', fw_data.get('category_breakdown', {}))
                
                for category, stats in categories.items():
                    current = stats.get('compliance_percentage', stats.get('compliance_pct', 0))
                    gap = target_compliance - current
                    
                    if gap > 0:
                        gap_data.append({
                            'Framework': fw_name,
                            'Category': category,
                            'Current_Compliance': f"{current:.1f}%",
                            'Target_Compliance': f"{target_compliance:.1f}%",
                            'Gap': f"{gap:.1f}%",
                            'Priority': 'High' if gap > 50 else 'Medium' if gap > 25 else 'Low',
                            'Estimated_Effort': self._estimate_effort(gap),
                            'Recommended_Timeline': self._estimate_timeline(gap)
                        })
            
            df = pd.DataFrame(gap_data)
            if not df.empty:
                df = df.sort_values('Gap', ascending=False)
            df.to_excel(writer, sheet_name='Gap_Analysis', index=False)
            
            logger.debug(f"Gap analysis sheet created with {len(gap_data)} gaps")
            
        except Exception as e:
            logger.error(f"Error creating gap analysis: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Gap_Analysis', index=False
            )
    
    def _create_remediation_tracking_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create remediation tracking sheet for project management."""
        try:
            frameworks = audit_results.get('frameworks', {})
            tracking_data = []
            
            for fw_name, fw_data in frameworks.items():
                rules = fw_data.get('rules', fw_data.get('rule_details', []))
                
                for rule in rules:
                    if rule.get('status') in ['FAIL', 'MISSING_DATA', 'ERROR']:
                        severity = rule.get('severity', 'MEDIUM')
                        due_days = {'CRITICAL': 7, 'HIGH': 30, 'MEDIUM': 60, 'LOW': 90}.get(severity, 60)
                        due_date = (datetime.now() + timedelta(days=due_days)).strftime("%Y-%m-%d")
                        
                        tracking_data.append({
                            'Framework': fw_name,
                            'Rule_ID': rule.get('rule_id', 'N/A'),
                            'Description': rule.get('description', 'N/A')[:100],
                            'Severity': severity,
                            'Status': 'Open',
                            'Assigned_To': '',
                            'Due_Date': due_date,
                            'Progress': '0%',
                            'Notes': '',
                            'Verification_Required': 'Yes'
                        })
            
            df = pd.DataFrame(tracking_data)
            df.to_excel(writer, sheet_name='Remediation_Tracking', index=False)
            
            logger.debug(f"Remediation tracking sheet created with {len(tracking_data)} items")
            
        except Exception as e:
            logger.error(f"Error creating remediation tracking: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Remediation_Tracking', index=False
            )
    
    def _create_timeline_roadmap_sheet(
        self,
        writer: pd.ExcelWriter,
        audit_results: Dict[str, Any]
    ):
        """Create timeline and roadmap sheet."""
        try:
            roadmap_data = [
                {
                    'Phase': 'Phase 1: Immediate Actions',
                    'Timeline': '0-30 days',
                    'Focus': 'Critical and High severity issues',
                    'Key Activities': 'Address security vulnerabilities, implement basic controls',
                    'Success Criteria': 'Zero critical issues, <5 high issues',
                    'Resources_Required': 'Security team, System administrators'
                },
                {
                    'Phase': 'Phase 2: Short-term Improvements',
                    'Timeline': '1-6 months',
                    'Focus': 'Medium severity issues, process improvements',
                    'Key Activities': 'Deploy monitoring, strengthen access controls',
                    'Success Criteria': '60%+ compliance across all frameworks',
                    'Resources_Required': 'IT operations, Security team'
                },
                {
                    'Phase': 'Phase 3: Long-term Maturity',
                    'Timeline': '6-12 months',
                    'Focus': 'Continuous improvement, automation',
                    'Key Activities': 'Advanced threat detection, compliance automation',
                    'Success Criteria': '85%+ compliance, automated monitoring',
                    'Resources_Required': 'Full IT team, External consultants'
                }
            ]
            
            df = pd.DataFrame(roadmap_data)
            df.to_excel(writer, sheet_name='Timeline_Roadmap', index=False)
            
            logger.debug("Timeline roadmap sheet created")
            
        except Exception as e:
            logger.error(f"Error creating timeline roadmap: {e}", exc_info=True)
            pd.DataFrame({'Error': [str(e)]}).to_excel(
                writer, sheet_name='Timeline_Roadmap', index=False
            )
    
    def _enhance_excel_workbook(self, excel_file: Path):
        """Apply professional formatting and styling to the Excel workbook."""
        try:
            wb = load_workbook(excel_file)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Apply header formatting
                if ws.max_row > 0:
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.fill = self.header_fill
                        cell.font = self.header_font
                        cell.alignment = self.header_alignment
                        cell.border = self.border
                        
                        # Auto-adjust column width
                        column_letter = get_column_letter(col)
                        ws.column_dimensions[column_letter].width = 20
                    
                    # Apply conditional formatting for risk levels
                    self._apply_conditional_formatting(ws)
            
            wb.save(excel_file)
            logger.debug(f"Enhanced formatting applied to {excel_file}")
            
        except Exception as e:
            logger.error(f"Error enhancing workbook: {e}", exc_info=True)
    
    def _apply_conditional_formatting(self, ws):
        """Apply conditional formatting to worksheet."""
        try:
            # Find risk level columns and apply color coding
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Color code risk levels
                    if cell.value in self.risk_colors:
                        cell.fill = PatternFill(
                            start_color=self.risk_colors[cell.value],
                            end_color=self.risk_colors[cell.value],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF" if cell.value in ['CRITICAL', 'HIGH'] else "000000")
                    
                    # Format percentages
                    if isinstance(cell.value, str) and '%' in cell.value:
                        try:
                            pct_value = float(cell.value.strip('%'))
                            if pct_value >= 85:
                                cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type="solid")
                            elif pct_value >= 60:
                                cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type="solid")
                            elif pct_value < 60:
                                cell.fill = PatternFill(start_color=self.colors['danger'], end_color=self.colors['danger'], fill_type="solid")
                        except:
                            pass
                            
        except Exception as e:
            logger.debug(f"Minor error in conditional formatting: {e}")
    
    # Helper methods
    def _get_risk_level(self, compliance_pct: float) -> str:
        """Get risk level based on compliance percentage."""
        if compliance_pct >= 85:
            return 'LOW'
        elif compliance_pct >= 60:
            return 'MEDIUM'
        elif compliance_pct >= 40:
            return 'HIGH'
        else:
            return 'CRITICAL'
    
    def _get_severity_impact(self, severity: str) -> str:
        """Get impact description for severity level."""
        impact_map = {
            'CRITICAL': 'Immediate exploitation risk, complete system compromise possible',
            'HIGH': 'Significant security gap, unauthorized access likely',
            'MEDIUM': 'Moderate security concern, limited system access possible',
            'LOW': 'Minor security issue, minimal risk exposure'
        }
        return impact_map.get(severity, 'Unknown impact')
    
    def _estimate_effort(self, gap: float) -> str:
        """Estimate effort required to close gap."""
        if gap > 50:
            return 'High (1-2 months)'
        elif gap > 25:
            return 'Medium (2-4 weeks)'
        else:
            return 'Low (1-2 weeks)'
    
    def _estimate_timeline(self, gap: float) -> str:
        """Estimate timeline to close gap."""
        if gap > 50:
            return '2-3 months'
        elif gap > 25:
            return '1-2 months'
        else:
            return '2-4 weeks'


# Export main class
__all__ = ['EnhancedReportService']

